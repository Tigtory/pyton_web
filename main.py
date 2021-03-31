from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import Updater, MessageHandler, Filters, CommandHandler
from openpyxl import load_workbook

TOKEN = '1630917550:AAE09IrOaZ6Pg7KgU3meDlZxu2xvfhv05zQ'
book = load_workbook('база данных.xlsx')
sheet = book['Лист1']
stickers_page = book['Стикеры']

def main():
    updater = Updater(token=TOKEN)    # объект, который ловит сообщения из телеграм
    dispatcher = updater.dispatcher
    handler = MessageHandler(Filters.all, do_echo)   # отфильтровываем сообщения: теперь устройство должно реагировать
    start_handler = CommandHandler('start', do_start)
    help_handler = CommandHandler('help', do_help)
    sticker_handler = MessageHandler(Filters.sticker, do_sticker)
    keyboard_handler = MessageHandler(Filters.text, do_something)


    dispatcher.add_handler(start_handler)
    dispatcher.add_handler(help_handler)
    dispatcher.add_handler(sticker_handler)
    dispatcher.add_handler(keyboard_handler)
    dispatcher.add_handler(handler)

    updater.start_polling()
    updater.idle()

def do_echo(update: Update, context):
    user = update.message.from_user.is_bot
    name = update.message.from_user.first_name
    if user == True:
        update.message.reply_text(text=f"Ты - бот! Уходи отсюда!!!")
    else:
        update.message.reply_text(text=f"ААААА! {name} что ты делаешь?")
        update.message.reply_text(text="Я не понимаю")

def do_start(update, context):
    keyboard = [
        ["1", "2", "3"],
        ["hello", "bye"]
    ]
    update.message.reply_text(text="Ты запустил меня человек", reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True))      # текст сообщения, который бот автоматически будет выдавать
def do_help(update, context):
    update.message.reply_text(text="что случилось? Всё ж было норм")

def do_something(update, context):
    text = update.message.text

    for row in range(2, stickers_page.max_row + 1):
        catch_phrase = stickers_page.cell(row=row, column=4).value
        print(catch_phrase)
        print(text)
        if catch_phrase in text:
            stickers_id = stickers_page.cell(row=row, column=3).value
            update.message.reply_sticker(stickers_id)
            return

    if text == "1":
        update.message.reply_text(text="Эта кнопка 1", reply_markup=ReplyKeyboardRemove())
    elif text == "2":
        update.message.reply_text(text="А это кнопка 2", reply_markup=ReplyKeyboardRemove())
        update.message.reply_sticker('CAACAgIAAxkBAAObYEimXjZXqlBA5MMrh0MGMjOu7AkAAlYsAALpVQUYUZfaCoweVn8eBA')
    elif text == "3":
        update.message.reply_text(text="Но вот эта кнопка 3", reply_markup=ReplyKeyboardRemove())

def do_sticker(update: Update, context):
    sticker_id = update.message.sticker.file_id
    update.message.reply_text(sticker_id)

main()