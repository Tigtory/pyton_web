from openpyxl import load_workbook


book = load_workbook('база данных.xlsx')
sheet = book['Лист1']
sheet['A1'].value = 1
stickers_page = book['Стикеры']
book.save('база данных.xlsx')

print(book.worksheets)
for i in range(1,4):
    print(stickers_page.cell(row=1, column=i).value, end='\t')